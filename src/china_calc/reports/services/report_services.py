from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from china_calc.finance.services.client_balance_service import ClientBalanceService


class ShipmentReportServices:
    def __init__(self, shipment, client):
        self.shipment = shipment
        self.client = client

    @staticmethod
    def append_payment_row(sheet, label, value, currency, fill_color, font_color):

        sheet.append([label, value])

        row = sheet.max_row
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for column in range(1, 3):
            cell = sheet.cell(row=row, column=column)
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        sheet.cell(row=row, column=1).font = Font(color=font_color, bold=True)

        value_cell = sheet.cell(row=row, column=2)
        value_cell.font = Font(color=font_color, bold=True)
        value_cell.number_format = f"#,##0.00 '{currency}'"

        sheet.row_dimensions[row].height = 22

    def build(self):
        items = list(
            self.shipment.items.filter(client=self.client)
            .select_related("client")
            .order_by("pk")
        )

        if not items:
            raise ValueError(
                "Невозможно скачать отчёт: у клиента нет товаров в этой поставке"
            )

        calculation_result = self.shipment.calculation_results.filter(
            is_actual=True
        ).first()

        if calculation_result is None:
            raise ValueError(
                "Невозможно скачать отчёт: у поставки нет актуального расчета"
            )

        client_result = (
            calculation_result.client_results.filter(client=self.client)
            .prefetch_related("item_results__item")
            .first()
        )

        if client_result is None:
            raise ValueError(
                "Невозможно скачать отчёт: для клиента нет результатов расчета"
            )

        item_results = {
            result.item_id: result for result in client_result.item_results.all()
        }
        balance = ClientBalanceService.calculate(
            shipment=self.shipment,
            client_result=client_result,
        )
        route_name = (
            f"{self.shipment.get_route_type_display()} ."
            f"{self.shipment.get_transport_type_display()}"
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Отчет клиента"

        sheet.append(
            [
                f"Отчет о поставке {self.shipment.number}",
            ]
        )
        sheet["A1"].font = Font(
            name="Carlito",
            size=16,
            bold=True,
        )

        sheet.append(
            [
                f"Клиент: {self.client.full_name}",
            ]
        )

        sheet["A2"].font = Font(
            name="Carlito",
            size=13,
            bold=True,
        )

        sheet.append([None])

        sheet.append(
            ["Маршрут", "Статус", "Способ расчета", "Тариф за 1 кг", "Тариф за 1 м3"]
        )

        sheet.append(
            [
                route_name,
                self.shipment.get_status_display(),
                self.shipment.get_logistic_calculation_type_display(),
                self.shipment.tariff_one_kg,
                self.shipment.tariff_one_m3,
            ]
        )

        tariff_currency = str(self.shipment.tariff_currency).replace('"', "")

        for column in range(4, 6):
            sheet.cell(
                row=sheet.max_row,
                column=column,
            ).number_format = f'#,##0.00 "{tariff_currency}"'

        sheet.append([None])

        sheet.append(
            [
                "Результат расчета клиента",
                "Сумма",
            ]
        )

        currency = str(calculation_result.final_currency)

        calculation_rows = [
            ("Стоимость товара", client_result.client_purchase_cost),
            ("Логистика", client_result.logistics_cost),
            ("Расходы", client_result.expenses_cost),
            ("Комиссия", client_result.buyer_commission_cost),
            ("Итого клиенту", client_result.client_price_cost),
        ]

        for label, value in calculation_rows:
            sheet.append([label, value])

            current_row = sheet.max_row
            value_cell = sheet.cell(row=current_row, column=2)
            value_cell.number_format = f'#,##0.00 "{currency}"'

            if label == "Итого клиенту":
                for column in range(1, 3):
                    cell = sheet.cell(row=current_row, column=column)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor="FFF2CC",
                    )

        sheet.append([])

        # Отдельный блок взаиморасчётов
        sheet.append(["Взаиморасчёты с клиентом", "Сумма"])

        payment_header_row = sheet.max_row

        for column in range(1, 3):
            cell = sheet.cell(row=payment_header_row, column=column)
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="4472C4",
            )
            cell.alignment = Alignment(vertical="center")

        sheet.row_dimensions[payment_header_row].height = 24

        paid_amount = balance["paid_amount"]
        debt_amount = balance["debt_amount"]
        overpayment_amount = balance["overpayment_amount"]

        # Оплачено — всегда зелёный
        self.append_payment_row(
            sheet=sheet,
            label="Оплачено",
            value=paid_amount,
            currency=currency,
            fill_color="E2F0D9",
            font_color="006100",
        )

        # Долг выделяем красным только при наличии долга
        self.append_payment_row(
            sheet=sheet,
            label="Осталось оплатить",
            value=debt_amount,
            currency=currency,
            fill_color="FCE4D6" if debt_amount > 0 else "F2F2F2",
            font_color="C00000" if debt_amount > 0 else "7F7F7F",
        )

        # Переплату выделяем синим только при её наличии
        self.append_payment_row(
            sheet=sheet,
            label="Переплата",
            value=overpayment_amount,
            currency=currency,
            fill_color="DDEBF7" if overpayment_amount > 0 else "F2F2F2",
            font_color="0070C0" if overpayment_amount > 0 else "7F7F7F",
        )

        sheet.append([None])

        sheet.append(
            [
                "Товар",
                "Трек-номер",
                "Количество",
                "Цена",
                "Вес, кг",
                "Объем, м3",
                "Стоимость товара",
                "Логистика",
                "Прямые расходы",
                "Общие расходы",
                "Итого",
                "Ссылка",
            ]
        )

        final_currency = str(calculation_result.final_currency)

        for item in items:
            item_result = item_results.get(item.pk)

            if item_result is None:
                raise ValueError(f"Для товара {item.pk} отсутствует результат расчета")

            sheet.append(
                [
                    item.name,
                    item.tracking_number,
                    item.quantity,
                    item.price,
                    item.weight,
                    item.volume,
                    item_result.client_purchase_cost,
                    item_result.logistics_cost,
                    item_result.direct_expenses_cost,
                    item_result.distributed_expenses_cost,
                    item_result.total_cost,
                    item.product_link,
                ]
            )

            price_currency = str(item.price_currency)

            sheet.cell(
                row=sheet.max_row,
                column=4,
            ).number_format = f'#,##0.00 "{price_currency}"'

            for column in range(7, 12):
                sheet.cell(
                    row=sheet.max_row,
                    column=column,
                ).number_format = f'#,##0.00 "{final_currency}"'

            if item.product_link:
                link_cell = sheet.cell(row=sheet.max_row, column=12)
                link_cell.hyperlink = item.product_link
                link_cell.style = "Hyperlink"

        column_width = {
            "A": 40,
            "B": 18,
            "C": 16,
            "D": 14,
            "E": 14,
            "F": 12,
            "G": 22,
            "H": 16,
            "I": 16,
            "J": 16,
            "K": 16,
            "L": 18,
            "M": 18,
        }

        for column, width in column_width.items():
            sheet.column_dimensions[column].width = width

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        return buffer
